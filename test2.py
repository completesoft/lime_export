import requests
import json
import base64
from openpyxl import load_workbook
import csv
from datetime import datetime
from string import ascii_uppercase


PREFIX_BASE = "#"
PREFIX_DATE = "$"
PREFIX_PLUS = "+"
PREFIX_PARSE = "%"
PREFIX_LIST = "*"


url = 'http://forms.product.in.ua/index.php/admin/remotecontrol'
login = 'admin'
password = 'ghjuhfvb'


# def jsonRequest(url, payload):
# 	headers = {'content-type': 'application/json'}
# 	r = requests.post(url, data=json.dumps(payload), headers = headers)
# 	return json.loads(r.text)


def json_request(j_url, j_payload):
    headers = {'content-type': 'application/json'}
    r = requests.post(j_url, data=json.dumps(j_payload), headers=headers)
    return r.json()


def get_session_key(m_url):
    payload = {'method': 'get_session_key', 'params': (login, password), 'id': 1}
    j = json_request(m_url, payload)
    return j["result"]


def release_session_key(m_url, key):
    payload = {"method": "release_session_key", "params": key, "id": 1}
    j = json_request(m_url, payload)
    return j["result"]


def export_responses(m_url, key, sid):
    payload = {"method": "export_responses", "params": (key, sid, "csv", "ru", "full"), "id": 1}
    j = json_request(m_url, payload)
    return base64.b64decode(j["result"])


def list_questions(m_url, key, sid):
    payload = {"method": "list_questions", "params": (key, sid), "id": 1}
    j = json_request(m_url, payload)
    return j["result"]


def date_format(datetime_string):
    try:
        dateformat = datetime.strptime(datetime_string, '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
    except ValueError:
        dateformat = datetime_string

    return dateformat


def to_xlsx(row_data):
    wb = load_workbook(filename='template2.xlsx')
    ws = wb.active
    ws['E2'] = date_format(row_data['submitdate'])
    ws['A2'] = row_data['Q00']
    ws['A6'] = row_data['Q01']
    ws['A9'] = date_format(row_data['Q02'])
    ws['B9'] = row_data['Q06']
    if row_data['Q07'] == 'Y':
        ws['D9'] = row_data['Q06']
    else:
        ws['D9'] = row_data['Q09']
    ws['A14'] = row_data['Q10']
    if (row_data['Q13']+row_data['Q14']) == "A1":
        ws['B14'] = "женат/замужем"
    else:
        ws['B14'] = "холост/не замужем"

    ws['D14'] = row_data['Q17']
    wb.save("sample.xlsx")


def to_xls_from_template(template, row_data, out_file):
    wb = load_workbook(filename=template)
    ws = wb.active

    for col_index in ascii_uppercase:
        for row_index in range(1, 100):
            ws_index = col_index + str(row_index)       # index xls cell
            cell_data = str(ws[ws_index].value)         # data xls cell
            template_col_index = cell_data[1:]          # data xls cell without prefix (#, $, ...)
            prefix = cell_data[0]
            # simple write
            if prefix == PREFIX_BASE:
                print(cell_data+" -> "+row_data[template_col_index])
                ws[ws_index] = row_data[template_col_index]

            # write with date format conversion
            if prefix == PREFIX_DATE:
                print(cell_data + " -> " + date_format(row_data[template_col_index]))
                ws[ws_index] = date_format(row_data[template_col_index])

            # write with parsing
            if prefix == PREFIX_PARSE:
                ex_left = template_col_index.split("=")[0]
                ex_right = template_col_index.split("=")[1]
                template_col_values = ex_right.split("|")
                template_col_k1 = template_col_values[0]
                template_col_v1 = template_col_values[1]
                template_col_k2 = template_col_values[2]
                template_col_v2 = template_col_values[3]
                template_col_kv = {template_col_k1: template_col_v1, template_col_k2: template_col_v2}
                ws[ws_index] = template_col_kv[row_data[ex_left]]

            # write with plus parsing
            if prefix == PREFIX_PLUS:
                ex_left = template_col_index.split("=")[0]
                ex_right = template_col_index.split("=")[1]
                template_col_index_sum1 = ex_left.split(",")[0]
                template_col_index_sum2 = ex_left.split(",")[1]
                template_col_values = ex_right.split("|")
                template_col_k1 = template_col_values[0]
                template_col_v1 = template_col_values[1]
                template_col_k2 = template_col_values[2]
                template_col_v2 = template_col_values[3]
                template_col_kv = {template_col_k1: template_col_v1, template_col_k2: template_col_v2}
                result_data = row_data[template_col_index_sum1]+row_data[template_col_index_sum2]
                ws[ws_index] = template_col_kv[result_data]

            if prefix == PREFIX_LIST:
                list_index = template_col_index.split(",")
                result_string = ""
                for list_item in list_index:
                    result_string = result_string + " " + date_format(row_data[list_item])
                ws[ws_index] = result_string


    try:
        wb.save(out_file)
    except PermissionError:
        print("!!!File is busy!!!")


def export_to_csv():

    session_key = get_session_key(url)

    questions_list = list_questions(url, session_key, "563799")

    print(questions_list)

    for question in questions_list:
        qid = question["id"]
        title = question["title"]
        question_text = question["question"]
        print(title + "    " + question_text)

    my_csv = export_responses(url, session_key, "563799").decode("utf-8")
    print(my_csv)
    f = open("export.csv", "w")
    f.write(my_csv)
    f.close()

    release_session_key(url, session_key)

export_to_csv()

reader = csv.DictReader(open("export.csv"))
for row in reader:
    if row['Q01'] != '':
        print(row['submitdate'] + "  " + row['Q00'])
        to_xls_from_template("template2.xlsx", row, row["id"]+"_"+row["Q01"]+".xlsx")







